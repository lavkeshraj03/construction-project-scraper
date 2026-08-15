"""app/exporters/excel.py — Excel export with 5 sheets.

Sheets:
1. Projects   — main client-facing sheet (frozen header, auto-filter, table)
2. Raw Data   — source traceability (project_id + raw JSON)
3. Sources    — where each record came from
4. Scrape Log — per-run per-source statistics
5. Summary    — aggregated counts

The file is always fully regenerated (not appended) to keep it clean.
"""

from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.config import settings
from app.processors.scorer import lead_priority
from app.utils.logger import get_logger

log = get_logger("exporter.excel")

# ─── Column definitions ────────────────────────────────────────────────────────

PROJECTS_COLUMNS: list[tuple[str, str]] = [
    # Client-required fields only
    ("Builder Name",             "builder_name"),
    ("Project Name",             "project_name"),
    ("Location",                 "location"),
    ("Project Value",            "project_value"),
    ("Decision Maker",           "decision_maker"),
    ("Mobile",                   "mobile"),
    ("Email",                    "email"),
    ("Architect",                "architect"),
    ("Contractor",               "contractor"),
    ("Builder / Architect / Contractor", "builder_architect_contractor"),
    ("Current Stage",            "current_stage"),
    ("Lead Score",               "lead_score"),
    ("Expected Order Value",     "expected_order_value"),
    ("Competition",              "competition"),
    ("Material Required",        "material_required"),
    ("Material Categories",      "material_categories"),
]

SOURCES_COLUMNS = [
    "Project ID", "Source", "Source URL", "Source Type", "Collected At", "Status", "Confidence",
]

RAW_COLUMNS = ["Project ID", "Source", "Source URL", "Raw Data"]

LOG_COLUMNS = [
    "Run ID", "Source", "Start Time", "End Time",
    "Records Found", "Records Added", "Duplicates", "Errors", "Status",
]

# ─── Styling ──────────────────────────────────────────────────────────────────

_HEADER_FILL  = PatternFill("solid", fgColor="1F3864")   # dark navy
_HEADER_FONT  = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
_ALT_FILL     = PatternFill("solid", fgColor="EEF2FF")   # light blue-grey
_NORMAL_FONT  = Font(name="Calibri", size=10)
_ALIGN_WRAP   = Alignment(wrap_text=True, vertical="top")
_ALIGN_CENTER = Alignment(horizontal="center", vertical="top")

_STAGE_COLOURS: dict[str, str] = {
    "Under Construction":    "C6EFCE",  # green
    "Advanced Construction": "A9D08E",
    "Near Completion":       "FFEB9C",  # yellow
    "Planning":              "DDEBF7",
    "Approved":              "BDD7EE",
    "Pre-Construction":      "FCE4D6",
    "Completed":             "D9D9D9",
    "Unknown":               "FFFFFF",
}

_PRIORITY_COLOURS: dict[str, str] = {
    "Very High": "C6EFCE",
    "High":      "FFEB9C",
    "Medium":    "FCE4D6",
    "Low":       "D9D9D9",
}


def _apply_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _ALIGN_CENTER


def _autosize_columns(ws, max_width: int = 50) -> None:
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, max_width)


def _to_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M")
    return str(val)


# ─── Sheet builders ────────────────────────────────────────────────────────────

def _build_projects_sheet(wb: openpyxl.Workbook, projects: list[Any]) -> None:
    ws = wb.active
    ws.title = "Projects"
    headers = [h for h, _ in PROJECTS_COLUMNS]
    _apply_header(ws, headers)

    for row_idx, proj in enumerate(projects, start=2):
        # Build row values
        row_vals: list[Any] = []
        for _, field in PROJECTS_COLUMNS:
            val = getattr(proj, field, None)
            row_vals.append(_to_str(val))

        ws.append(row_vals)

        # Alternate row shading
        fill = _ALT_FILL if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

        for col_idx, (_, field) in enumerate(PROJECTS_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = _NORMAL_FONT
            cell.alignment = _ALIGN_WRAP

            # Stage colour override
            if field == "current_stage":
                stage = _to_str(getattr(proj, "current_stage", ""))
                colour = _STAGE_COLOURS.get(stage, "FFFFFF")
                cell.fill = PatternFill("solid", fgColor=colour)
            else:
                cell.fill = fill

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    if projects:
        ws.auto_filter.ref = ws.dimensions

    # Excel table
    if len(projects) > 0:
        last_col = get_column_letter(len(PROJECTS_COLUMNS))
        last_row = 1 + len(projects)
        tbl = Table(displayName="ProjectsTable", ref=f"A1:{last_col}{last_row}")
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tbl.tableStyleInfo = style
        ws.add_table(tbl)

    _autosize_columns(ws)


def _build_raw_sheet(wb: openpyxl.Workbook, projects: list[Any]) -> None:
    ws = wb.create_sheet("Raw Data")
    _apply_header(ws, RAW_COLUMNS)
    for proj in projects:
        ws.append([
            _to_str(proj.project_id),
            _to_str(proj.source),
            _to_str(proj.source_url),
            _to_str(proj.raw_data),
        ])
        ws.cell(row=ws.max_row, column=4).alignment = _ALIGN_WRAP
    _autosize_columns(ws, max_width=80)


def _build_sources_sheet(wb: openpyxl.Workbook, projects: list[Any]) -> None:
    ws = wb.create_sheet("Sources")
    _apply_header(ws, SOURCES_COLUMNS)
    for proj in projects:
        ws.append([
            _to_str(proj.project_id),
            _to_str(proj.source),
            _to_str(proj.source_url),
            _to_str(proj.source),            # Source Type = source name
            _to_str(proj.created_at),
            _to_str(proj.duplicate_status),
            _to_str(proj.confidence_score),
        ])
    _autosize_columns(ws)


def _build_log_sheet(wb: openpyxl.Workbook, runs: list[Any]) -> None:
    ws = wb.create_sheet("Scrape Log")
    _apply_header(ws, LOG_COLUMNS)
    for run in runs:
        ws.append([
            _to_str(run.run_id),
            _to_str(run.source),
            _to_str(run.start_time),
            _to_str(run.end_time),
            run.records_found or 0,
            run.records_added or 0,
            run.duplicates or 0,
            run.errors or 0,
            _to_str(run.status),
        ])
    _autosize_columns(ws)


def _build_summary_sheet(wb: openpyxl.Workbook, projects: list[Any], runs: list[Any]) -> None:
    ws = wb.create_sheet("Summary")
    ws.title = "Summary"

    title_font = Font(bold=True, size=14, name="Calibri")
    label_font = Font(bold=True, name="Calibri")
    value_font = Font(name="Calibri")

    ws["A1"] = "Construction Project Scraper — Summary"
    ws["A1"].font = title_font
    ws.merge_cells("A1:C1")

    stats: list[tuple[str, Any]] = [
        ("", ""),
        ("Total Projects",        len(projects)),
        ("Unique Projects",       sum(1 for p in projects if getattr(p, "duplicate_status", "") != "duplicate")),
        ("Duplicates",            sum(1 for p in projects if getattr(p, "duplicate_status", "") == "duplicate")),
        ("", ""),
        ("Construction Stage Breakdown", ""),
        ("Under Construction",   sum(1 for p in projects if getattr(p, "current_stage", "") == "Under Construction")),
        ("Advanced Construction", sum(1 for p in projects if getattr(p, "current_stage", "") == "Advanced Construction")),
        ("Near Completion",      sum(1 for p in projects if getattr(p, "current_stage", "") == "Near Completion")),
        ("Planning",             sum(1 for p in projects if getattr(p, "current_stage", "") == "Planning")),
        ("Completed",            sum(1 for p in projects if getattr(p, "current_stage", "") == "Completed")),
        ("Unknown Stage",        sum(1 for p in projects if getattr(p, "current_stage", "") == "Unknown")),
        ("", ""),
        ("Lead Priority Breakdown", ""),
        ("Very High Priority",   sum(1 for p in projects if lead_priority(int(getattr(p, "lead_score", 0) or 0)) == "Very High")),
        ("High Priority",        sum(1 for p in projects if lead_priority(int(getattr(p, "lead_score", 0) or 0)) == "High")),
        ("Medium Priority",      sum(1 for p in projects if lead_priority(int(getattr(p, "lead_score", 0) or 0)) == "Medium")),
        ("Low Priority",         sum(1 for p in projects if lead_priority(int(getattr(p, "lead_score", 0) or 0)) == "Low")),
        ("", ""),
        ("Contact Information", ""),
        ("Decision Maker Found", sum(1 for p in projects if getattr(p, "decision_maker", "Not Found") not in ("Not Found", "", None))),
        ("Mobile Found",         sum(1 for p in projects if getattr(p, "mobile", "Not Found") not in ("Not Found", "", None))),
        ("Email Found",          sum(1 for p in projects if getattr(p, "email", "Not Found") not in ("Not Found", "", None))),
        ("Architect Found",      sum(1 for p in projects if getattr(p, "architect", "Not Found") not in ("Not Found", "", None))),
        ("Contractor Found",     sum(1 for p in projects if getattr(p, "contractor", "Not Found") not in ("Not Found", "", None))),
        ("", ""),
        ("Material Categories", ""),
        ("Granite Projects",     sum(1 for p in projects if "Granite" in str(getattr(p, "material_categories", "")))),
        ("Marble Projects",      sum(1 for p in projects if "Marble" in str(getattr(p, "material_categories", "")))),
        ("Quartz Projects",      sum(1 for p in projects if "Quartz" in str(getattr(p, "material_categories", "")))),
        ("Kota Projects",        sum(1 for p in projects if "Kota" in str(getattr(p, "material_categories", "")))),
        ("Material Relevant",    sum(1 for p in projects if getattr(p, "material_required", "") == "Yes")),
        ("", ""),
        ("Sources", ""),
    ]

    # Source breakdown
    source_counts: dict[str, int] = {}
    for p in projects:
        src = getattr(p, "source", "Unknown")
        source_counts[src] = source_counts.get(src, 0) + 1
    for src, count in sorted(source_counts.items()):
        stats.append((f"  {src}", count))

    stats += [
        ("", ""),
        ("Scraper Runs", len(runs)),
        ("Generated At", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
    ]

    for row_data in stats:
        label, value = row_data
        ws.append([label, value])
        row = ws.max_row
        ws.cell(row=row, column=1).font = label_font if label else value_font
        ws.cell(row=row, column=2).font = value_font

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18


# ─── Main export function ─────────────────────────────────────────────────────

def export_to_excel(
    projects: list[Any],
    runs: list[Any],
    output_path: Path | None = None,
) -> Path:
    """
    Generate the Excel workbook with all 5 sheets.
    Returns the path to the generated file.
    """
    path = Path(output_path or settings.OUTPUT_PATH)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    _build_projects_sheet(wb, projects)
    _build_raw_sheet(wb, projects)
    _build_sources_sheet(wb, projects)
    _build_log_sheet(wb, runs)
    _build_summary_sheet(wb, projects, runs)

    wb.save(path)
    log.info("Excel exported: %s (%d projects, %d scrape runs)", path, len(projects), len(runs))
    return path
